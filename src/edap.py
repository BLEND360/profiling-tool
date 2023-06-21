import pandas as pd
import numpy as np
from scipy import stats
from sklearn.tree import DecisionTreeClassifier
from sklearn.model_selection import GridSearchCV
from openpyxl import *
import openpyxl
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.drawing.text import ParagraphProperties, CharacterProperties
from openpyxl.styles.borders import Border
from openpyxl.formatting.rule import ColorScaleRule
import scipy.cluster.hierarchy as sch
from scipy.spatial.distance import pdist
from openpyxl.styles import PatternFill, Alignment
from tqdm import tqdm
import win32com.client as win32
from openpyxl.worksheet.datavalidation import DataValidation
from sklearn.preprocessing import MinMaxScaler
from sklearn.preprocessing import StandardScaler
import os
import warnings
import ast
from numpy.lib.stride_tricks import sliding_window_view
from kmodes.kprototypes import KPrototypes
import re


class ExcelReport:
    def __init__(self, file_name, null_percentage_to_drop_column=None, dependant_target_variable=None):
        self.__file_name = file_name
        self.__per_to_drop = null_percentage_to_drop_column
        self.__dependant_target_variable = dependant_target_variable
        self.df = self.__load_data(file_name, dependant_target_variable)
        # self.col_to_drop = None
        self.__df_org = None
        self.__df_original = None
        self.__df_after_missing_impute = None
        self.__wb = Workbook()
        self.__ws = self.__wb['Sheet']
        self.__ws.title = 'Buckets'
        self.__s1 = self.__wb['Buckets']

    def __load_data(self, file_path, dep_var):
        """
        Load the data from a given file path and set the index column

        :param file_path: Path to the dataset file
        :param dep_var: Dependent target variable column
        :return df: Dataframe which contains the loaded data with index column set
        """
        self.df = pd.read_csv(file_path, dtype={dep_var: 'object'}, low_memory=False)
        if self.__dependant_target_variable is not None:
            target_col = self.df.pop(self.__dependant_target_variable)
            self.df = pd.concat([self.df, target_col], axis=1)
            if self.df[self.__dependant_target_variable].isnull().sum() != 0:
                warnings.warn(f"Target column have null values check and try again", UserWarning)
        unique_columns = self.df.columns[self.df.nunique() == self.df.shape[0]]
        if unique_columns != 0:
            self.df = self.df.set_index(unique_columns[0])
        return self.df

    def __outlier_per_col(self, col):
        """
        Calculate the outlier percentage of a column

        :param col: Column Name
        :return outlier_per: Outlier percentage of the column
        """
        q1 = self.df[col].quantile(0.25)
        q3 = self.df[col].quantile(0.75)
        iqr = q3 - q1

        # Kolmogorov-Smirnov test to find the distribution of the data
        dist_name, p = stats.normaltest(self.df[col])[0], stats.normaltest(self.df[col])[1]

        # if p > 0.05 then the data is normally distributed
        # if p <= 0.05 then the data is not normally is distributed
        if p <= 0.05:
            lower_bound = q1 - 1.5 * iqr
            upper_bound = q3 + 1.5 * iqr
            outlier_df = self.df[(self.df[col] < lower_bound) | (self.df[col] > upper_bound)]
            outlier_per = (len(outlier_df) / len(self.df[col])) * 100
        else:
            z_score = np.abs(self.df[col] - self.df[col].mean()) / self.df[col].std()
            outlier_df = self.df[(z_score > 3)]
            outlier_per = len(outlier_df) / len(self.df[col]) * 100
        return outlier_per

    def __calc_woe_iv(self, col, y0, y1):
        """
        Calculate the IV value for a column

        :param col: Column Name
        :param y0: y0
        :param y1: y1
        :return: IV value of a column
        """
        if col == self.__dependant_target_variable:
            return '-'
        elif self.df[col].nunique() == 0 or self.df[col].nunique() == 1:
            return 0
        else:
            all_cols = self.df[self.__dependant_target_variable].unique()
            ta = pd.crosstab(self.df[col], self.df[self.__dependant_target_variable], normalize='columns')
            if len(self.df[self.__dependant_target_variable]) != len(ta.columns.tolist()):
                ta = ta.reindex(columns=all_cols, fill_value=0)
            woe_iv = (ta.assign(woe=lambda dfx: np.log((dfx[1] + (0.5 / y1)) / (dfx[0] + (0.5 / y0)))).assign(iv=lambda dfx: sum(dfx['woe'] * (dfx[1] - dfx[0]))))
            return woe_iv['iv'].unique()

    def generate_report(self):
        """
        Generated an excl which contains the data summary for every column

        :return df: Dataframe which contains the loaded data
        """
        summary_df = self.df.isna().sum().reset_index().rename(columns={'index': 'variable', 0: 'null'})
        summary_df['%null'] = (100 * summary_df['null'] / len(self.df)).round(2)
        summary_df = summary_df.merge(self.df.dtypes.reset_index().rename(columns={'index': 'variable', 0: 'type'}), on='variable')
        summary_df = summary_df.drop(columns=['null'])
        if self.__per_to_drop is not None:
            summary_df = summary_df.drop(summary_df[summary_df['%null'] > self.__per_to_drop].index)
        # null_percentage = self.df.isnull().sum() / self.df.shape[0] * 100
        # col_to_drop = null_percentage[null_percentage > 90].keys()
        # self.df = self.df.drop(col_to_drop, axis=1)
        df_numeric = self.df.select_dtypes(exclude='object')
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            summary_df['outlier%'] = summary_df[summary_df['variable'].isin(df_numeric.columns)].apply(lambda x: self.__outlier_per_col(x['variable']), axis=1)
        summary_df = summary_df.merge((self.df.select_dtypes(exclude=['object']).nunique() / self.df.select_dtypes(exclude=['object']).count() * 100).reset_index().rename(columns={'index': 'variable', 0: 'unique%'}).round(2), on='variable', how='left').round(2)
        summary_df = summary_df.merge(self.df.mean(numeric_only=True).reset_index().rename(columns={'index': 'variable', 0: 'mean'}).round(2), on='variable', how='left')
        summary_df = summary_df.merge(self.df.std(numeric_only=True).reset_index().rename(columns={'index': 'variable', 0: 'standard deviation'}).round(2), on='variable', how='left')
        summary_df = (summary_df.merge(self.df.var(numeric_only=True).reset_index().rename(columns={'index': 'variable', 0: 'variance'}), on='variable', how='left').assign(variance=lambda x: x['variance'].apply(lambda y: "{:.2f}".format(y))))
        summary_df = summary_df.merge(self.df.skew(numeric_only=True).reset_index().rename(columns={'index': 'variable', 0: 'skewness'}).round(2), on='variable', how='left')
        summary_df = summary_df.merge(self.df.kurt(numeric_only=True).reset_index().rename(columns={'index': 'variable', 0: 'kurtosis'}).round(2), on='variable', how='left')
        summary_df = summary_df.merge(self.df.min(numeric_only=True).reset_index().rename(columns={'index': 'variable', 0: 'min'}), on='variable', how='left')
        summary_df = summary_df.merge(self.df.max(numeric_only=True).reset_index().rename(columns={'index': 'variable', 0: 'max'}), on='variable', how='left')
        summary_df['range'] = summary_df['max'] - summary_df['min']
        summary_df = summary_df.merge((self.df.describe().loc['75%'].T - self.df.describe().loc['25%'].T).reset_index().rename(columns={'index': 'variable', 0: 'iqr'}), on='variable', how='left')
        summary_df = summary_df.merge(self.df.median(numeric_only=True).reset_index().rename(columns={'index': 'variable', 0: 'median'}), on='variable', how='left')
        summary_df = summary_df.merge(self.df.select_dtypes(include=['object']).mode().iloc[0].reset_index().rename(columns={'index': 'variable', 0: 'mode'}), on='variable', how='left')
        summary_df = summary_df.merge(self.df.select_dtypes(include=['object']).nunique().reset_index().rename(columns={'index': 'variable', 0: 'distinct count'}), on='variable', how='left')
        summary_df['user action missing'] = np.where(summary_df['type'] == 'object', 'mode', 'mean')
        summary_df['user action outlier'] = np.where(summary_df['type'] != 'object', 'capping', '-')
        summary_df['user action type'] = np.where(summary_df['type'] == 'object', 'nominal', '-')
        summary_df['user action ordinal'] = np.where(summary_df['type'] == 'object', '', '-')
        summary_df['Min-Max'] = np.where(summary_df['type'] != 'object', 'N', '-')
        summary_df['Standardization'] = np.where(summary_df['type'] != 'object', 'N', '-')
        summary_df['Log'] = np.where(summary_df['type'] != 'object', '', '-')
        summary_df['Lead'] = 0
        summary_df['Lag'] = 0
        summary_df['Adstock'] = np.where(summary_df['type'] != 'object', '', '-')
        if self.__dependant_target_variable is not None:
            self.df[self.__dependant_target_variable] = self.df[self.__dependant_target_variable].astype('int64')
            y0 = self.df[self.__dependant_target_variable].value_counts()[0]
            y1 = self.df[self.__dependant_target_variable].value_counts()[1]
            summary_df = summary_df.merge(self.df.apply(lambda x: self.__calc_woe_iv(x.name, y0, y1)).T.reset_index().rename(columns={'index': 'variable', 0: 'IV'}), on='variable', how='left')
            df_target = summary_df.iloc[-1, :]
            df_target = pd.DataFrame(data=df_target).T
            summary_df = summary_df.iloc[:-1, :]
            summary_df = summary_df.sort_values(by=['IV'], ascending=False)
            summary_df['IV'] = summary_df['IV'].apply(lambda x: round(x, 2))
            summary_df = pd.concat([summary_df, df_target], axis=0)
        summary_df.to_excel(self.__file_name + '_eda_report_2.xlsx', index=False, freeze_panes=(1, 1))

        wb = openpyxl.load_workbook(self.__file_name + '_eda_report_2.xlsx')
        ws = wb.active
        col_ranges = ['R', 'S', 'T', 'V', 'W', 'Y', 'Z']
        valid_options = [
            '"mean, median, mode, linear interpolation, polynomial interpolation, drop row, drop column, forward fill, backward fill"',
            '"capping, remove, mean, median"',
            '"nominal, ordinal"',
            '"Y, N"',
            '"Y, N"',
            '"0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12"',
            '"0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12"']
        for i in range(len(col_ranges)):
            rule = DataValidation(type='list', formula1=valid_options[i], allow_blank=True)
            rule.error = 'Your entry is not valid.'
            rule.errorTitle = 'Invalid Entry'
            rule.prompt = 'Please select from thr list.'
            rule.promptTitle = 'Select Option'
            ws.add_data_validation(rule)
            column_letter = col_ranges[i]
            column_range = f'{column_letter}2:{column_letter}{ws.max_row}'
            for cell in ws[column_range]:
                if cell[0].value != '-':
                    rule.add(cell[0])
        prompt_log = 'multiplier|addition'
        dv_log = DataValidation(prompt=prompt_log)
        dv_range = 'X1'
        ws.add_data_validation(dv_log)
        dv_log.add(dv_range)
        prompt_adstock = 'adstock|lag'
        dv_adstock = DataValidation(prompt=prompt_adstock)
        dv_range_ad = 'AA1'
        ws.add_data_validation(dv_adstock)
        dv_adstock.add(dv_range_ad)
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 20
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        wb.save(self.__file_name + '_eda_report_2.xlsx')
        # return self.df

    # def impute(self,col,mth):#mean, median, mode, ffill,bfill
    # missing value treatments
    def __mean_imputation(self, col):
        """
        Mean imputation for missing values

        :param col: Column Name
        """
        self.df[col].fillna(round(self.df[col].mean(), 2), inplace=True)
        # return self.df

    def __median_imputation(self, col):
        """
        Median imputation for missing values

        :param col: Column Name
        """
        median = self.df[col].median()
        self.df[col].fillna(round(median, 2), inplace=True)
        # return self.df

    def __drop_rows(self, col):
        """
        Drop rows with missing values

        :param col: Column Name
        """
        self.df.dropna(subset=[col], inplace=True)
        # return self.df

    def __drop_column(self, col):
        """
        Drop column

        :param col: Column Name
        """
        self.df.drop(col, axis=1, inplace=True)
        # return self.df

    def __mode_imputation(self, col):
        """
        Mode imputation for missing values

        :param col: Column Name
        """
        mode = self.df[col].mode()[0]
        self.df[col].fillna(mode, inplace=True)
        # return self.df

    def __arbitrary_val(self, col, val):
        """
        Fill the missing values with a specific value

        :param col: Column Name
        :param val: Value to impute in place of missing values
        """
        self.df[col].fillna(val, inplace=True)
        # return self.df

    def __linear_interpolate(self, col):
        """
        Linear interpolation for missing values

        :param col: Column Name
        """
        self.df[col].interpolate(method='linear', inplace=True)
        # return self.df

    def __polynomial_interpolate(self, col):
        """
        Polynomial interpolation for missing values

        :param col: Column Name
        """
        self.df[col].interpolate(method='polynomial', order=2, inplace=True)
        # return self.df

    def __interpolate_padding_forward(self, col):
        """
        Interpolation with padding forward for missing values

        :param col: Column Name
        """
        self.df[col].fillna(method='ffill', inplace=True)
        # return self.df

    def __interpolate_padding_backward(self, col):
        """
        Interpolation with padding backward for missing values

        :param col: Column Name
        """
        self.df[col].fillna(method='bfill', inplace=True)
        # return self.df

    # outlier treatments
    def __remove_outliers(self, col):
        """
        Remove outliers for a specific column

        :param col: Column Name
        """
        dist_name, p = stats.normaltest(self.df[col])[0], stats.normaltest(self.df[col])[1]
        if p <= 0.05:
            q1 = self.df[col].quantile(0.25)
            q3 = self.df[col].quantile(0.75)
            iqr = q3 - q1
            lower_bound = q1 - 1.5 * iqr
            upper_bound = q3 + 1.5 * iqr
            self.df = self.df[(self.df[col] >= lower_bound) & (self.df[col] <= upper_bound)]
        else:
            z_score = np.abs(self.df[col] - self.df[col].mean()) / self.df[col].std()
            self.df = self.df[(z_score < 3)]
        return self.df

    def __mean_outlier(self, col):
        """
        Mean imputation for outliers

        :param col: Column Name
        """
        dist_name, p = stats.normaltest(self.df[col])[0], stats.normaltest(self.df[col])[1]
        if p <= 0.05:
            q1 = self.df[col].quantile(0.25)
            q3 = self.df[col].quantile(0.75)
            iqr = q3 - q1
            lower_bound = q1 - 1.5 * iqr
            upper_bound = q3 + 1.5 * iqr
            self.df[col][self.df[col] < lower_bound] = self.df[col].mean()
            self.df[col][self.df[col] > upper_bound] = self.df[col].mean()
        else:
            z_score = np.abs(self.df[col] - self.df[col].mean()) / self.df[col].std()
            self.df.loc[z_score > 3, col] = self.df[col].mean()
        return self.df

    def __median_outlier(self, col):
        """
        Median imputation for outliers

        :param col: Column Name
        """
        dist_name, p = stats.normaltest(self.df[col])[0], stats.normaltest(self.df[col])[1]
        if p <= 0.05:
            q1 = self.df[col].quantile(0.25)
            q3 = self.df[col].quantile(0.75)
            iqr = q3 - q1
            lower_bound = q1 - 1.5 * iqr
            upper_bound = q3 + 1.5 * iqr
            self.df[col][self.df[col] < lower_bound] = self.df[col].median()
            self.df[col][self.df[col] > upper_bound] = self.df[col].median()
        else:
            z_score = np.abs(self.df[col] - self.df[col].mean()) / self.df[col].std()
            self.df.loc[z_score > 3, col] = self.df[col].median()
        return self.df

    def __outlier_capping(self, col):
        """
        Capping the outliers with upper limit and lower limit

        :param col: Column Name
        """
        dist_name, p = stats.normaltest(self.df[col])[0], stats.normaltest(self.df[col])[1]
        if p <= 0.05:
            # mu = self.df[col].mean()
            # sigma = self.df[col].std()
            # scaled_data = (self.df[col] - mu) / sigma
            # q1 = scaled_data.quantile(0.25)
            # q3 = scaled_data.quantile(0.75)
            # iqr = q3 - q1
            # upper_bound = q3 + 1.5 * iqr
            # lower_bound = q1 - 1.5 * iqr
            # capped_data = np.where(scaled_data >= upper_bound, upper_bound, np.where(scaled_data <= lower_bound, lower_bound, scaled_data))
            # self.df[col] = capped_data * sigma + mu
            q1 = self.df[col].quantile(0.25)
            q3 = self.df[col].quantile(0.75)
            iqr = q3-q1
            lower_bound = q1-1.5*iqr
            upper_bound = q1+1.5*iqr
            self.df[col] = np.where(self.df[col] >= upper_bound, upper_bound, np.where(self.df[col] <= lower_bound, lower_bound, self.df[col]))
        else:
            # mu = self.df[col].mean()
            # sigma = self.df[col].std()
            # scaled_data = (self.df[col] - mu) / sigma
            # upper_limit = scaled_data.mean() + (3 * scaled_data.std())
            # lower_limit = scaled_data.mean() - (3 * scaled_data.std())
            # capped_data = np.where(scaled_data >= upper_limit, upper_limit, np.where(scaled_data <= lower_limit, lower_limit, scaled_data))
            # self.df[col] = capped_data * sigma + mu
            upper_limit = self.df[col].mean() + (3 * self.df[col].std())
            lower_limit = self.df[col].mean() - (3 * self.df[col].std())
            self.df[col] = np.where(self.df[col] >= upper_limit, upper_limit, np.where(self.df[col] <= lower_limit, lower_limit, self.df[col]))
        return self.df

    def __perform_treatment_missing(self, col_name, treatments):
        """
        Perform missing value treatment on a column in the dataframe

        :param col_name: Column Name
        :param treatments: Type of missing value treatment to be applied
        :return: The treated column or the original column if no treatment was performed
        """
        if treatments == 'mean':
            self.__mean_imputation(col_name)
        elif treatments == 'median':
            self.__median_imputation(col_name)
        elif treatments == 'drop row':
            self.__drop_rows(col_name)
        elif treatments == 'drop column':
            self.__drop_column(col_name)
        # elif all(word in treatments for word in ['interpolation', 'linear']):
        elif treatments == 'linear interpolation':
            self.__linear_interpolate(col_name)
        # elif all(word in treatments for word in ['interpolation', 'polynomial']):
        elif treatments == 'polynomial interpolation':
            self.__polynomial_interpolate(col_name)
        # elif all(word in treatments for word in ['interpolation', 'padding', 'forward']):
        elif treatments == 'forward fill':
            self.__interpolate_padding_forward(col_name)
        # elif all(word in treatments for word in ['interpolation', 'padding', 'backward']):
        elif treatments == 'backward fill':
            self.__interpolate_padding_backward(col_name)
        elif treatments == 'mode':
            self.__mode_imputation(col_name)
        else:
            return self.df[col_name]

    def __perform_treatment_outlier(self, col_name, treatments):
        """
        Perform outlier value treatment on a column in the dataframe

        :param col_name: Column Name
        :param treatments: The type of outlier value treatment to be applied
        :return: The treated column or the original column if no treatment was performed
        """
        if treatments == 'remove':
            self.__remove_outliers(col_name)
        elif treatments == 'mean':
            self.__mean_outlier(col_name)
        elif treatments == 'median':
            self.__median_imputation(col_name)
        elif treatments == 'capping':
            self.__outlier_capping(col_name)
        else:
            return self.df[col_name]

    # this function generated a new csv file after transforming the data
    def __transform(self, treatment_file=None):
        """
        Perform all the necessary treatments for missing values and outliers for all the columns

        :param treatment_file: Path to the generated Excel file
        :return df: Dataframe which contains all the treated data
        """
        print('transformation start')
        self.__df_original = self.df.copy()
        if self.__per_to_drop is not None:
            null_percentage = self.df.isnull().sum() / self.df.shape[0] * 100
            col_to_drop = null_percentage[null_percentage > self.__per_to_drop].keys()
            self.df = self.df.drop(col_to_drop, axis=1)

        cols_with_one_unique = self.df.columns[self.df.nunique() == 1]
        self.df.drop(cols_with_one_unique, axis=1, inplace=True)
        # null_percentage = self.df.isnull().sum() / self.df.shape[0] * 100
        # self.col_to_drop = null_percentage[null_percentage > self.per_to_drop].keys()
        # cols_with_one_unique = self.df.columns[self.df.nunique() == 1]
        # self.col_to_drop.extend(list(set(cols_with_one_unique)-set(self.col_to_drop)))
        # self.df.drop(self.col_to_drop, axis=1, inplace=True)
        if treatment_file is None:
            print('transformation end')
            print()
            return self.df
        else:
            df1 = pd.read_excel(treatment_file)
            # df1 = pd.read_excel(self.file_name + '_eda_report_2.xlsx')
            treatments_dict_missing = dict(zip(df1['variable'], df1['user action missing']))
            treatments_dict_outlier = dict(zip(df1['variable'], df1['user action outlier']))

            # df_original = self.df.copy()

            self.df.apply(lambda col: self.__perform_treatment_missing(col.name, treatments_dict_missing.get(col.name, None)), axis=0)
            self.__df_after_missing_impute = self.df.copy()
            self.df.apply(lambda col: self.__perform_treatment_outlier(col.name, treatments_dict_outlier.get(col.name, None)), axis=0)

            if self.__dependant_target_variable is not None:
                df2 = df1.copy()
                df2 = df2[df2['IV'] != '-']
                df2['IV'] = pd.to_numeric(df2['IV'])
                df2 = df2.sort_values(by='IV', ascending=False)
                col_names_by_iv = df2['variable'].tolist()

                col_names_by_iv.append(str(self.__dependant_target_variable))
                common_cols = [col for col in col_names_by_iv if col in self.df.columns]
                self.df = self.df[common_cols]

            # lead transformation
            treatment_lead = dict(zip(df1['variable'], df1['Lead']))
            treatment_lead = {k: v for k, v in treatment_lead.items() if v not in [0, '0']}
            if len(treatment_lead) != 0:
                for k, v in treatment_lead.items():
                    if k in self.df.columns:
                        self.df[k] = self.df[k].shift(-v)
                        self.df[k] = self.df[k].fillna(0)

            # lag transformation
            treatment_lag = dict(zip(df1['variable'], df1['Lag']))
            treatment_lag = {k: v for k, v in treatment_lag.items() if v not in [0, '0']}
            if len(treatment_lag) != 0:
                for k, v in treatment_lag.items():
                    self.df[k] = self.df[k].shift(v)
                    self.df[k] = self.df[k].fillna(0)

            # adstock transformation
            treatment_adstock = dict(zip(df1['variable'], df1['Adstock']))
            treatment_adstock = {k: v for k, v in treatment_adstock.items() if v not in ['-', np.nan]}
            if len(treatment_adstock) != 0:
                for k, v in treatment_adstock.items():
                    if type(v) == float or type(v) == int:
                        # make changes so that the value is not negative
                        if v > 1:
                            warnings.warn(f"Value of adstock is greater than 1", UserWarning)
                        else:
                            x = 0
                            self.df[k] = [x := x * v + i for i in self.df[k]]
                    elif '|' in v:
                        adstock, lag = v.split('|')

                        def check_type(my_string):
                            try:
                                my_float = float(my_string)
                                if my_float.is_integer():
                                    val = 'int'
                                else:
                                    val = 'float'
                            except ValueError:
                                val = 'not valid'
                            return val

                        if check_type(adstock) != 'not valid':
                            adstock = float(adstock)
                            if adstock > 1 or adstock < 0:
                                warnings.warn(f"Value of adstock should be in between 0, 1", UserWarning)
                        else:
                            warnings.warn(f"Check the value of adstock", UserWarning)

                        if check_type(lag) == 'int':
                            lag = int(lag)
                        else:
                            warnings.warn(f"Value of lag should be an integer", UserWarning)

                        # x = 0
                        # new = []
                        # count = 0
                        # for i in self.df[k]:
                        #     x = x * adstock + i
                        #     new.append(x)
                        #     count += 1
                        #     if lag == count:
                        #         break
                        # weights = []
                        # for j in range(lag):
                        #     weights.append(adstock ** (j+1))
                        # for i in range(lag, len(self.df)):
                        #     temp = self.df[k].iloc[i]
                        #     for j in range(lag):
                        #         temp += weights[j] * self.df[k].iloc[i-j-1]
                        #     new.append(temp)
                        # self.df[k] = new

                        my_list = self.df[k].tolist()
                        new_list = [0]*lag + my_list
                        v = sliding_window_view(new_list, lag+1)
                        weights = []
                        for j in range(lag):
                            weights.append(adstock ** (j+1))
                        weights = [1]+weights
                        weights_arr = np.tile(np.flip(weights), (len(self.df), 1))
                        result = np.dot(weights_arr, v.T)
                        first_array = result[0].tolist()
                        self.df[k] = first_array

                    else:
                        warnings.warn(f"Give the input in the correct format", UserWarning)

            # min-max normalization & standardization
            cols_to_normalize = df1[df1['Min-Max'] == 'Y']['variable']
            cols_to_standardize = df1[df1['Standardization'] == 'Y']['variable']
            if list(np.intersect1d(cols_to_normalize, cols_to_standardize)):
                warnings.warn(f"Cannot perform both min-max normalization and standardization on a column", UserWarning)
            else:
                if len(cols_to_normalize) != 0:
                    scaler = MinMaxScaler()
                    self.df[cols_to_normalize] = scaler.fit_transform(self.df[cols_to_normalize])
                if len(cols_to_standardize) != 0:
                    std_scaler = StandardScaler()
                    self.df[cols_to_standardize] = std_scaler.fit_transform(self.df[cols_to_standardize])

            # log transformation
            treatment_log = dict(zip(df1['variable'], df1['Log']))
            treatment_log = {k: v for k, v in treatment_log.items() if v not in ['-', np.nan]}
            for k, v in treatment_log.items():
                multiplier, addition = v.split('|')
                self.df[k] = np.log(self.df[k] * float(multiplier) + float(addition) + (-(self.df[k].min()) if self.df[k].min() < 0 else 0))

            # self.df.to_csv(self.__file_name + '_transformed_data_2.csv', index=False)
            print('transformation end')
            print()
            return self.df, self.__df_original, self.__df_after_missing_impute

    def show_transformed_data(self):
        """
        This function shows the dataframe after all the transformations are done

        :return: Dataframe after all transformations are done
        """
        return self.df

    def code_generator(self, treatment_file=None):
        """
        This function generates a code when run it will give the dataframe after all the transformations are done

        :param treatment_file: Path to the summary file
        :return: Python code to get the dataframe after all transformations are done
        """
        # a, b, c = self.transform(treatment_file=treatment_file)
        b = self.__df_original
        c = self.__df_after_missing_impute
        df_num = self.df.select_dtypes(exclude=['object'])
        if treatment_file is not None:
            treat_df = pd.read_excel(treatment_file)
            treatments_dict_missing = dict(zip(treat_df['variable'], treat_df['user action missing']))
            treatments_dict_outlier = dict(zip(treat_df['variable'], treat_df['user action outlier']))
            cols_to_drop = list(filter(lambda x: x not in self.df.columns, b.columns))
            treatments_dict_missing = dict(filter(lambda x: x[0] in self.df.columns, treatments_dict_missing.items()))
            treatments_dict_outlier = dict(filter(lambda x: x[0] in self.df.columns, treatments_dict_outlier.items()))
            treatments_dict_outlier = dict(filter(lambda x: x[0] in df_num.columns, treatments_dict_outlier.items()))
            code = f"df = pd.read_csv('{self.__file_name}')\n"
            code += f"df = df.drop({cols_to_drop}, axis=1)"
            code += "\n"
            for k, v in treatments_dict_missing.items():
                if v == 'mean':
                    impute_mean = round(b[k].mean(), 2)
                    code += f"df['{k}'] = df['{k}'].fillna({impute_mean})\n"
                elif v == 'median':
                    impute_median = round(b[k].median(), 2)
                    code += f"df['{k}'] = df['{k}'].fillna({impute_median})\n"
                elif v == 'mode':
                    impute_mode = b[k].mode()[0]
                    code += f"df['{k}'] = df['{k}'].fillna({repr(impute_mode)})\n"
                elif v == 'forward fill':
                    code += f"df['{k}'] = df['{k}'].fillna(method='ffill')\n"
                elif v == 'backward fill':
                    code += f"df['{k}'] = df['{k}'].fillna(method='bfill')\n"
                elif v == 'linear interpolation':
                    code += f"df['{k}'] = df['{k}'].interpolate(method='linear')\n"
                elif v == 'polynomial interpolation':
                    code += f"df['{k}'] = df['{k}'].interpolate(method='polynomial, order=2)'\n"
                elif v == 'drop row':
                    code += f"df = df.dropna(subset=['{k}'])\n"
                elif v == 'drop column':
                    code += f"df = df.drop('{k}', axis=1)\n"
            code += "\n"

            for k, v in treatments_dict_outlier.items():
                dist_name, p = stats.normaltest(c[k])[0], stats.normaltest(c[k])[1]
                mean_out = c[k].mean()
                median_out = c[k].median()
                std_out = c[k].std()
                if p <= 0.05:
                    q1 = c[k].quantile(0.25)
                    q3 = c[k].quantile(0.75)
                    iqr = q3 - q1
                    lower_bound = q1 - 1.5 * iqr
                    upper_bound = q3 + 1.5 * iqr
                    if v == 'remove':
                        code += f"df = df[(df['{k}'] >= {lower_bound}) & (df['{k}'] <= {upper_bound})]\n"
                    elif v == 'mean':
                        code += f"df['{k}'][df['{k}'] < {lower_bound}] = {mean_out}\n"
                        code += f"df['{k}'][df['{k}'] > {upper_bound}] = {mean_out}\n"
                    elif v == 'median':
                        code += f"df['{k}'][df['{k}'] < {lower_bound}] = {median_out}\n"
                        code += f"df['{k}'][df['{k}'] > {upper_bound}] = {median_out}\n"
                    elif v == 'capping':
                        code += f"df['{k}'] = np.where(df['{k}'] >= {upper_bound}, {upper_bound}, np.where(df['{k}'] <= {lower_bound}, {lower_bound}, df['{k}']))\n"
                else:
                    z_score = np.abs(c[k] - mean_out / std_out)
                    upper_limit = mean_out + 3 * std_out
                    lower_limit = mean_out - 3 * std_out
                    if v == 'remove':
                        code += f"df = df[({z_score} < 3)]\n"
                    elif v == 'mean':
                        code += f"df.loc[{z_score} > 3, '{k}'] = {mean_out}\n"
                    elif v == 'median':
                        code += f"df.loc[{z_score} > 3, '{k}'] = {median_out}\n"
                    elif v == 'capping':
                        code += f"df['{k}'] = np.where(df['{k}'] >= {upper_limit}, {upper_limit}, np.where(df['{k}'] <= {lower_limit}, {lower_limit}, df['{k}']))\n"
            code += "\n"

            treatment_lead = dict(zip(treat_df['variable'], treat_df['Lead']))
            treatment_lead = {k: v for k, v in treatment_lead.items() if v not in [0, '0']}
            if len(treatment_lead) != 0:
                for k, v in treatment_lead.items():
                    code += f"df['{k}'] = df['{k}'].shift(-{v})\n"
                    code += f"df['{k}'] = df['{k}'].fillna(0)\n"
            code += "\n"

            treatment_lag = dict(zip(treat_df['variable'], treat_df['Lag']))
            treatment_lag = {k: v for k, v in treatment_lag.items() if v not in [0, '0']}
            if len(treatment_lag) != 0:
                for k, v in treatment_lag.items():
                    code += f"df['{k}'] = df['{k}'].shift({v})\n"
                    code += f"df['{k}'] = df['{k}'].fillna(0)\n"
            code += "\n"

            treatment_adstock = dict(zip(treat_df['variable'], treat_df['Adstock']))
            treatment_adstock = {k: v for k, v in treatment_adstock.items() if v not in ['-', np.nan]}
            if len(treatment_adstock) != 0:
                for k, v in treatment_adstock.items():
                    if type(v) == float or type(v) == int:
                        if v <= 1:
                            code += f"x = 0\n"
                            code += f"df['{k}'] = [x := x * {v} + i for i in df['{k}']]\n"
                        elif '|' in v:
                            adstock, lag = v.split('|')

                            def check_type(my_string):
                                try:
                                    my_float = float(my_string)
                                    if my_float.is_integer():
                                        val = 'int'
                                    else:
                                        val = 'float'
                                except ValueError:
                                    val = 'not found'
                                return val

                            if check_type(lag) == 'int':
                                lag = int(lag)

                            code += f"my_list = df[{k}].tolist()\n"
                            code += f"new_list = [0] * {lag} + my_list\n"
                            code += f"v = sliding_window_view(new_list, {lag} + 1\n)"
                            code += f"weights = []\n"
                            code += f"for j in range({lag}):\n"
                            code += f"    weights.append({adstock} ** (j + 1))\n"
                            code += f"weights = [1] + weights\n"
                            code += f"weights_arr = np.tile(np.flip(weights), (len(df), 1))\n"
                            code += f"result = np.dor(weights_arr, v.T)\n"
                            code += f"first_array = result[0].tolist()\n"
                            code += f"df[{k}] = first_array\n"
            code += "\n"

            cols_to_normalize = treat_df[treat_df['Min-Max'] == 'Y']['variable']
            cols_to_standardize = treat_df[treat_df['Standardization'] == 'Y']['variable']
            if len(cols_to_normalize) != 0:
                code += f"scaler = MinMaxScaler()\n"
                for col in cols_to_normalize:
                    code += f"df['{col}'] = scaler.fit_transform(df['{col}'])\n"
            code += "\n"
            if len(cols_to_standardize) != 0:
                code += f"std_scaler = StandardScaler()\n"
                for col in cols_to_standardize:
                    code += f"df['{col}'] = std_scaler.fit_transform(df['{col}'])\n"
            code += "\n"

            treatment_log = dict(zip(treat_df['variable'], treat_df['Log']))
            treatment_log = {k: v for k, v in treatment_log.items() if v not in ['-', np.nan]}
            if len(treatment_log) != 0:
                for k, v in treatment_log.items():
                    multiplier, addition = v.split('|')
                    code += f"df['{k}'] = np.log(df['{k}'] * {multiplier} + {addition} + (-(df['{k}'].min()) if df['{k}'].min() < 0 else 0))\n"
            code += "\n"

            print(code)
        else:
            print('No treatments done')

    def __tree_based_bin_data(self, df, column_name, dependant_target_variable, depth_of_tree):
        """
        Perform tree based binning for a numerical column

        :param df: Dataframe
        :param column_name: Column Name
        :param dependant_target_variable: Dependant target variable
        :param depth_of_tree: Depth of the tree
        :return tree_based_binned_data: Dataframe which contains the binned data
        """
        df2 = self.__df_org.copy()
        df2 = df2.loc[df2[column_name].notnull()]
        x = df2[column_name].values.reshape(-1, 1)
        y = df2[dependant_target_variable].values
        params = {'max_depth': range(2, depth_of_tree + 1), 'min_samples_split': [2, 3, 5, 10], 'min_samples_leaf': [int(np.ceil(0.05 * len(x)))]}
        clf = DecisionTreeClassifier()
        g_search = GridSearchCV(clf, param_grid=params, scoring='accuracy')
        g_search.fit(x, y)
        best_clf = g_search.best_estimator_
        bin_edges = best_clf.tree_.threshold
        bin_edges = sorted(set(bin_edges[bin_edges != -2]))
        tree_based_binned_data = self.__value_bin_data(df, column_name, bin_edges)
        # print(column_name)
        return tree_based_binned_data

    def __decile_bin_data(self, df, col, no_of_bins):
        """
        Perform decile binning for a numerical column

        :param df: Dataframe
        :param col: Column Name
        :param no_of_bins: Maximum number of bins allowed
        :return decile_binned_data: Dataframe which contains the binned data
        """
        decile_binned_data = pd.qcut(df[col], no_of_bins, duplicates='drop')
        return decile_binned_data

    def __value_bin_data(self, df, col, no_of_bins):
        """
        Perform value based binning for a numerical column

        :param df: Dataframe
        :param col: Column Name
        :param no_of_bins: Maximum number of bins allowed
        :return value_binned_data: Dataframe which contains the binned data
        """
        # bins = [df[col].min()] + bins + [df[col].max()]
        # value_binned_data = pd.cut(df[col], bins, include_lowest=True, duplicates='drop')
        value_binned_data = pd.cut(df[col], no_of_bins, duplicates='drop')
        return value_binned_data

    def __col_bin_sumar(self, col, bin_df_1):
        """
        Generated the binned summary data for a column separately

        :param col: Column Name
        :param bin_df_1: Dataframe which contains all the binned data
        :return df_new: Dataframe which contains all the summary of binned data
        """
        unique_bin_edges = bin_df_1[col].unique()
        df_new = pd.DataFrame({"column_name": [col] * len(unique_bin_edges), "bin_ranges": unique_bin_edges})
        df_new = df_new.sort_values(by='bin_ranges')
        df_new = df_new.reset_index(drop=True)
        # df_new = df_new.merge((bin_df_1[col].value_counts() / len(bin_df_1) * 100).reset_index().rename(columns={'index': 'bin_ranges', col: 'count%'}).sort_values(by='bin_ranges').reset_index(drop=True), on='bin_ranges').round(2)
        df_new = df_new.merge((bin_df_1[col].value_counts() / len(bin_df_1) * 100).reset_index().rename(columns={col: 'bin_ranges', 'count': 'count%'}).sort_values(by='bin_ranges').reset_index(drop=True), on='bin_ranges').round(2)
        if self.__dependant_target_variable is not None:
            df_new = df_new.merge(bin_df_1.groupby(col).target.sum().reset_index().rename(columns={col: 'bin_ranges', 'target': 'Event'}), on='bin_ranges')
            df_new = df_new.merge(bin_df_1.groupby(col).target.mean().reset_index().rename(columns={col: 'bin_ranges', 'target': 'Mean_DV'}), on='bin_ranges')

            # while dividing two columns it will divide based on index, so it may come as NaN if the index do not match
            # so use .item() to not face this issue
            df_new['Index'] = (100 * df_new['Mean_DV'] / bin_df_1.target.mean()).round()
        return df_new

    def __create_binned_data(self, func='decile', dep_var=None, depth=None, no_of_bins=10):
        """
        Calculated bins for each numerical column in a pandas Dataframe and their summary with plots

        :param func: Which type of binning to perform on the data
        :param dep_var: Dependent target column
        :param depth: Depth of the tree if tree based binning
        :param no_of_bins: Maximum no of bins allowed if decile based binning and value based binning

        :return: An Excel file which contains the summary of each binned column and their respective charts
        """
        print('numerical binning start')
        self.__df_org = self.df.copy()
        if dep_var is not None:
            self.__df_org[dep_var] = self.__df_org[dep_var].astype('int64')
            self.__df_org = self.__df_org.rename(columns={dep_var: 'target'})
            dep_var = 'target'
            df_num = self.__df_org.select_dtypes(include=[np.number]).drop(dep_var, axis=1)
            if func == 'tree':
                tqdm.pandas(dynamic_ncols=True, position=0)
                bin_df = df_num.progress_apply(lambda x: self.__tree_based_bin_data(df_num, x.name, dep_var, depth))

                bin_df = bin_df.applymap(lambda x: 'NA' if pd.isnull(x) else x)
                bin_df = bin_df.astype('category')

                cols_with_one_unique_bin = bin_df.columns[bin_df.nunique() == 1]
                bin_df.drop(cols_with_one_unique_bin, axis=1, inplace=True)
            elif func == 'decile':
                tqdm.pandas(dynamic_ncols=True, position=0)
                bin_df = df_num.progress_apply(lambda x: self.__decile_bin_data(df_num, x.name, no_of_bins))

                bin_df = bin_df.applymap(lambda x: 'NA' if pd.isnull(x) else x)
                bin_df = bin_df.astype('category')

                cols_with_one_unique_bin = bin_df.columns[bin_df.nunique() == 1]
                bin_df.drop(cols_with_one_unique_bin, axis=1, inplace=True)
            else:
                tqdm.pandas(dynamic_ncols=True, position=0)
                bin_df = df_num.progress_apply(lambda x: self.__value_bin_data(df_num, x.name, no_of_bins))

                bin_df = bin_df.applymap(lambda x: 'NA' if pd.isnull(x) else x)
                bin_df = bin_df.astype('category')

                cols_with_one_unique_bin = bin_df.columns[bin_df.nunique() == 1]
                bin_df.drop(cols_with_one_unique_bin, axis=1, inplace=True)

            bin_df_1 = pd.concat([bin_df, self.__df_org[dep_var]], axis=1)
        else:
            df_num = self.__df_org.select_dtypes(include=[np.number])
            if func == 'decile':
                tqdm.pandas(dynamic_ncols=True, position=0)
                bin_df = df_num.progress_apply(lambda x: self.__decile_bin_data(df_num, x.name, no_of_bins))

                bin_df = bin_df.applymap(lambda x: 'NA' if pd.isnull(x) else x)
                bin_df = bin_df.astype('category')

                cols_with_one_unique_bin = bin_df.columns[bin_df.nunique() == 1]
                bin_df.drop(cols_with_one_unique_bin, axis=1, inplace=True)
            else:
                tqdm.pandas(dynamic_ncols=True, position=0)
                bin_df = df_num.progress_apply(lambda x: self.__value_bin_data(df_num, x.name, no_of_bins))

                bin_df = bin_df.applymap(lambda x: 'NA' if pd.isnull(x) else x)
                bin_df = bin_df.astype('category')

                cols_with_one_unique_bin = bin_df.columns[bin_df.nunique() == 1]
                bin_df.drop(cols_with_one_unique_bin, axis=1, inplace=True)

            bin_df_1 = bin_df.copy()
        # bin_df_1.rename(columns={dep_var: 'target'})
        print('numerical binning end')
        print()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Numerical Columns'

        # headers = ['column_name', 'bin_ranges', 'value_counts', 'Event', 'Mean_DV', 'Index']
        # self.write_summary_and_charts_to_excel(bin_df, self.col_bin_sumar, bin_df_1, wb, ws, 'bin_ranges', 7, 'H', header_vales=headers)

        row_num = 1
        if dep_var is not None:
            headers = ['column_name', 'bin_ranges', 'value_counts', 'Event', 'Mean_DV', 'Index']
            ws.append(headers)
            fill_color = PatternFill(start_color='89CFF0', end_color='89CFF0', fill_type='solid')
            for i in range(1, 7):
                ws.cell(row=1, column=i).border = openpyxl.styles.borders.Border(bottom=openpyxl.styles.borders.Side(style='thick'))
                ws.cell(row=1, column=i).fill = fill_color
        else:
            headers = ['column_name', 'bin_ranges', 'value_counts']
            ws.append(headers)
            fill_color = PatternFill(start_color='89CFF0', end_color='89CFF0', fill_type='solid')
            for i in range(1, 4):
                ws.cell(row=1, column=i).border = openpyxl.styles.borders.Border(bottom=openpyxl.styles.borders.Side(style='thick'))
                ws.cell(row=1, column=i).fill = fill_color

        for col in bin_df.columns:
            df_new = self.__col_bin_sumar(col, bin_df_1)
            df_new['bin_ranges'] = df_new['bin_ranges'].astype(str)
            start_row = row_num + 1
            for row in df_new.iterrows():
                row_num += 1
                for col_num, value in enumerate(row[1]):
                    ws.cell(row=row_num, column=col_num + 1, value=value)
                    ws.row_dimensions[row_num].height = 40
            end_row = row_num
            if dep_var is not None:
                for i in range(1, 7):
                    ws.cell(row=row_num, column=i).border = openpyxl.styles.borders.Border(bottom=openpyxl.styles.borders.Side(style='thick'))
            else:
                for i in range(1, 4):
                    ws.cell(row=row_num, column=i).border = openpyxl.styles.borders.Border(bottom=openpyxl.styles.borders.Side(style='thick'))

            if dep_var is not None:
                combo_chart = BarChart()
                combo_chart.y_axis.majorGridlines = None
                line_chart = LineChart()
                line_chart.y_axis.majorGridlines = None
                combo_chart.y_axis.title = 'Count%'
                combo_chart.x_axis.title = 'Bin Ranges'
                line_chart.y_axis.title = 'Index'
                combo_chart.title = 'Distribution and Index of ' + col
                combo_chart.style = 10
                no_of_rows = end_row - start_row + 1
                combo_chart.height = min(1.5 * no_of_rows, 10)
                combo_chart.width = 10
                line_chart.y_axis.crosses = "max"
                line_chart.y_axis.axId = 200

                value_counts = Reference(ws, min_col=3, min_row=row_num - len(df_new), max_row=row_num)
                bin_ranges = Reference(ws, min_col=2, min_row=row_num - len(df_new) + 1, max_row=row_num)
                index = Reference(ws, min_col=6, min_row=row_num - len(df_new), max_row=row_num)
                combo_chart.add_data(value_counts, titles_from_data=True)
                combo_chart.set_categories(bin_ranges)
                line_chart.add_data(index, titles_from_data=True)
                line_chart.set_categories(bin_ranges)

                combo_chart += line_chart
                combo_chart.legend = None
            else:
                combo_chart = BarChart()
                combo_chart.y_axis.majorGridlines = None
                combo_chart.y_axis.title = 'Count%'
                combo_chart.x_axis.title = 'Bin Ranges'
                combo_chart.title = 'Distribution and Index of ' + col
                combo_chart.style = 10
                no_of_rows = end_row - start_row + 1
                combo_chart.height = min(1.5 * no_of_rows, 10)
                combo_chart.width = 10

                value_counts = Reference(ws, min_col=3, min_row=row_num - len(df_new), max_row=row_num)
                bin_ranges = Reference(ws, min_col=2, min_row=row_num - len(df_new) + 1, max_row=row_num)
                combo_chart.add_data(value_counts, titles_from_data=True)
                combo_chart.set_categories(bin_ranges)

                combo_chart.legend = None

            def set_chart_title_size(chart, size):
                paraprops = ParagraphProperties()
                paraprops.defRPr = CharacterProperties(sz=size)
                for para in chart.title.tx.rich.paragraphs:
                    para.pPr = paraprops
            set_chart_title_size(combo_chart, size=1000)

            ws.add_chart(combo_chart, "H" + str(start_row))

            if dep_var is not None:
                x = ws.max_row
                rule = ColorScaleRule(start_type='min',
                                      start_color='00FF0000',
                                      mid_type='percentile',
                                      mid_value=50,
                                      mid_color='00FFFF00',  # Yellow
                                      end_type='max',
                                      end_color='0000FF00')
                ws.conditional_formatting.add(f"F2:F{x + 1}", rule)
        if dep_var is not None:
            ws["P2"].value = "Index: Index tells you how a bin is performing compared to the entire dataset."
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 15
        wb.save("profiled_data.xlsx")

    def __woe_iv(self, df, column_name, dependant_target_variable, no_of_bins):
        """
        Calculates the Weight of Evidence (WOE) and Information Value (IV) of a given column in a pandas DataFrame.

        :param df: DataFrame
        :param column_name: Column Name
        :param dependant_target_variable: Dependent variable column
        :param no_of_bins: Number of bins for grouping the data

        :return data: Series containing the binned values for the column with WOE applied.
        """

        y0 = df[dependant_target_variable].value_counts()[0]
        y1 = df[dependant_target_variable].value_counts()[1]
        if df[column_name].nunique() < 10:
            data = pd.Series(pd.factorize(df[column_name])[0] + 1, index=df.index).rename('{}'.format(column_name)).apply(lambda x: f'bin{x}')
        else:
            df_woe_iv = (pd.crosstab(df[column_name], df[dependant_target_variable], normalize='columns').assign(woe=lambda dfx: np.log((dfx[1] + (0.5 / y1)) / (dfx[0] + (0.5 / y0)))).assign(iv=lambda dfx: (dfx['woe'] * (dfx[1] - dfx[0]))))
            woe_map = df_woe_iv['woe'].to_dict()
            woe_col = df[column_name].map(woe_map)
            data = pd.qcut(woe_col, no_of_bins, duplicates='drop')
            n = data.nunique()
            labels = [f'bin{i}' for i in range(1, n + 1)]
            data = data.cat.rename_categories(labels)
            sizes = data.value_counts(normalize=True)
            min_size = 0.05
            while sizes.min() < min_size and no_of_bins > 1:
                no_of_bins -= 1
                data = pd.qcut(woe_col, q=no_of_bins, duplicates='drop')
                if data.nunique() != data.cat.categories.nunique():
                    continue
                n = data.nunique()
                labels = [f'bin{i}' for i in range(1, n + 1)]
                data = data.cat.rename_categories(labels)
                sizes = data.value_counts(normalize=True)
        return data

    def __naive_cat_bin(self, df, col, max_thre=10, min_thre=5, tolerence=2, flag='ignore'):
        """
        A naive approach for binning nominal categorical columns

        :param df: Data Frame
        :param col: Column name
        :param max_thre: Maximum Threshold
        :param min_thre: Minimum threshold
        :param tolerence: Tolerance
        :param flag: Flag
        :return: Column with binned data
        """
        value_counts = df[col].value_counts()
        total_values = len(df)
        count_percentages = (value_counts / total_values) * 100
        unique_values_df = pd.DataFrame({'Category': value_counts.index, 'Count Percentage': count_percentages})
        count_per = list(unique_values_df['Count Percentage'])

        final_ini = []
        for i in count_per:
            if i >= min_thre:
                final_ini.append(i)
        a = [x for x in count_per if x not in final_ini]

        total_bins = int(100 / max_thre)
        ava_bins = len(final_ini)
        ava_bin_per = sum(final_ini)
        bin_req = total_bins - ava_bins
        bin_req_per = 100 - ava_bin_per

        if flag == 'error' and bin_req > 0 and (bin_req_per / bin_req) > max_thre:
            print(f"Binning for {col} is not possible with given parameters.")
            return

        step = False
        while step == False:
            if bin_req > 0:
                if (bin_req_per / bin_req) > min_thre:
                    step = True
                else:
                    bin_req -= 1
            else:
                step = True

        final_ini = [[x] for x in final_ini]

        if bin_req > 0:
            target_sum = bin_req_per / bin_req
        else:
            target_sum = bin_req_per
            tolerence = 0

        final = []
        current_sum = 0.0
        start_index = len(a) - 1
        values = []
        while start_index >= 0:
            current_sum += a[start_index]
            values.append(a[start_index])
            if current_sum < target_sum - tolerence:
                start_index -= 1
            else:
                final.append(values)
                values = []
                start_index -= 1
                current_sum = 0.0
        final.append(values)
        final = final[::-1]
        final = [sublist for sublist in final if sublist]
        final_b = final_ini + final

        final = [final_b[0]]
        for subarr in final_b[1:]:
            if sum(subarr) < (min_thre - tolerence):
                final[-1].extend(subarr)
            else:
                final.append(subarr)

        table = dict(zip(unique_values_df['Category'], unique_values_df['Count Percentage']))
        new_final = [sublist.copy() for sublist in final]
        for i in range(len(new_final)):
            for j in range(len(new_final[i])):
                new_final[i][j] = next((k for k, v in table.items() if v == new_final[i][j]), None)
                table.pop(new_final[i][j], None)
        k = len(new_final)
        bin_labels = [f'bin{i}' for i in range(1, k + 1)]
        bin_mapping = {value: bin_labels[i] for i, sublist in enumerate(new_final) for value in sublist}
        bin_mapping[np.nan] = 'binNA'
        # df[col] = df[col].apply(lambda x: bin_mapping.get(x, x))
        return df[col].apply(lambda x: bin_mapping.get(x, x))

    def __naive_ord_cat_bin(self, df, col, ordinal_mapping, max_thre=10, min_thre=5, tolerence=2):
        """
        A naive approach to bin ordinal categorical columns

        :param df: Data Frame
        :param col: Column name
        :param ordinal_mapping: Order of the unique values in the column
        :param max_thre: Maximum threshold
        :param min_thre: Minimum threshold
        :param tolerence: Tolerance
        :return: Column with binned data
        """
        value_counts = df[col].value_counts()
        total_values = len(df)
        count_percentages = (value_counts / total_values) * 100
        unique_values_df = pd.DataFrame({'Category': value_counts.index, 'Count Percentage': count_percentages})
        unique_values_df = unique_values_df.iloc[unique_values_df.Category.map(ordinal_mapping).argsort()]
        count_per = list(unique_values_df['Count Percentage'])
        final_ini = [[x] for x in count_per]
        final = [final_ini[0]]
        for subarr in final_ini[1:]:
            if sum(subarr) < (min_thre - tolerence):
                final[-1].extend(subarr)
            else:
                final.append(subarr)
        total_bins = int(100 / max_thre)
        req_iter = len(final) - total_bins
        if req_iter > 0:
            for i in range(req_iter):
                min_sum = float('inf')
                idx1 = 0
                idx2 = 0
                for i in range(len(final) - 1, 0, -1):
                    combined_sum = sum(final[i]) + sum(final[i - 1])
                    if combined_sum < min_sum:
                        min_sum = combined_sum
                        idx1 = i - 1
                        idx2 = i
                combined_subarray = final[idx1] + final[idx2]
                final = [combined_subarray if idx == idx1 else subarray for idx, subarray in
                         enumerate(final[:idx2])] + final[idx2 + 1:]

        table = dict(zip(unique_values_df['Category'], unique_values_df['Count Percentage']))
        new_final = [sublist.copy() for sublist in final]
        for i in range(len(new_final)):
            for j in range(len(new_final[i])):
                new_final[i][j] = next((k for k, v in table.items() if v == new_final[i][j]), None)
                table.pop(new_final[i][j], None)

        k = len(new_final)
        bin_labels = [f'bin{i}' for i in range(1, k + 1)]
        bin_mapping = {value: bin_labels[i] for i, sublist in enumerate(new_final) for value in sublist}
        bin_mapping[np.nan] = 'binNA'
        return df[col].apply(lambda x: bin_mapping.get(x, x))

    def __hierarchical_clustering(self, df, column_name, ordinal_dict=None, min_cluster_size=0.05, max_clusters=10):
        """
        Hierarchical Clustering to perform binning on an ordinal categorical column

        :param df: Dataframe
        :param column_name: Column Name
        :param ordinal_dict: Order of the unique values in the column. If None, ordinal_dict will be created using sorted unique values of the column.
        :param min_cluster_size: Minimum size of clusters as a fraction of total population. Default is 0.05.
        :param max_clusters: Maximum number of clusters allowed. Default is 10.

        :return: Dataframe containing the binned values for the columns
        """
        df1 = df.copy()
        df1 = df1.loc[df1[column_name].notnull()]
        if ordinal_dict is None:
            ordinal_dict = {val: i + 1 for i, val in enumerate(sorted(df1[column_name].unique()))}
        mapped_data = df1[column_name].map(ordinal_dict)
        distances = pdist(mapped_data.values.reshape(-1, 1), metric='euclidean')

        linkage_matrix = sch.linkage(distances, method='ward')

        n_clusters = []
        unique_values = []
        layer = len(np.unique(linkage_matrix[:, 2]))
        total_population = len(df1)
        while True:
            cluster_labels = sch.fcluster(linkage_matrix, layer, criterion='maxclust')
            cluster_sizes = np.array([np.sum(cluster_labels == c) for c in np.unique(cluster_labels)])
            if np.min(cluster_sizes) >= min_cluster_size * total_population or layer == 1:
                n_clusters.append(np.unique(cluster_labels).size)
                unique_values.append([np.unique(df1[column_name].values[cluster_labels == c]) for c in np.unique(cluster_labels)])
                if n_clusters[-1] <= max_clusters:
                    break
                else:
                    layer -= 1
            else:
                layer -= 1

        # df[col_name] = cluster_labels
        # df[col_name] = df[col_name].apply(lambda x: f"cluster {x}")
        null_values = df[df[column_name].isnull()].copy()
        df = df.dropna(subset=[column_name])
        df.loc[:, column_name] = cluster_labels
        df.loc[:, column_name] = df.loc[:, column_name].apply(lambda x: f"cluster {x}")
        df = pd.concat([df, null_values], axis=0).sort_index()
        return df

    def __kprototype(self, df, col, ordinal_mapping=None, max_clusters=10):
        """
        KPrototypes based approach to perform binning on ordinal categorical columns

        :param df: Dataframe
        :param col: Column Name
        :param ordinal_mapping: Order of the unique values in a column
        :param max_clusters: Maximum number of clusters (bins) allowed

        :return: Dataframe containing the binned values for the columns
        """
        if ordinal_mapping is None:
            ordinal_mapping = {val: i + 1 for i, val in enumerate(sorted(df[col].unique()))}

        # if df[col].nunique() < 10:
        #     max_clusters = df[col].nunique()

        if df[col].isnull().sum() > 0:
            max_clusters -= 1
        old_data = list(map(str, df[col]))
        data = df.loc[df[col].notnull()]
        data = list(data[col])
        numeric_labels = np.array([ordinal_mapping[category] for category in data])
        labels = np.column_stack((numeric_labels, data))
        n_clusters = max_clusters
        kp = KPrototypes(n_clusters=n_clusters)
        clusters = kp.fit_predict(labels, categorical=[1])

        cluster_unique_values = {}
        for unique_value, cluster in zip(data, clusters):
            if cluster not in cluster_unique_values:
                cluster_unique_values[cluster] = [unique_value]
            else:
                if unique_value not in cluster_unique_values[cluster]:
                    cluster_unique_values[cluster].append(unique_value)
        cluster_unique_values['NA'] = ['nan']

        reverse_mapping = {value: key for key, values in cluster_unique_values.items() for value in values}
        df[col] = ['bin' + str(reverse_mapping[value]) if value in reverse_mapping else value for value in old_data]
        return df

    def __col_bin_sumar_cat(self, df_cat, col, binned_df_1):
        """
        Calculate the bin ranges, values in each bin, percentage of values, mean DV and index for each bin of a column in a pandas Dataframe

        :param df_cat: Dataframe containing the original data
        :param col: Column Name
        :param binned_df_1: Dataframe containing the binned data

        :return df_new: Dataframe containing the above calculated values
        """
        unique_values_in_bins = df_cat.groupby(binned_df_1[col])[col].unique().apply(list)
        unique_values_in_bins = unique_values_in_bins.rename_axis('bin').reset_index()
        # unique_bin_ranges = pd.Categorical(binned_df_1[col].unique()).sort_values(ascending=True)
        unique_bin_ranges = pd.Categorical(binned_df_1[col].unique())
        uni = binned_df_1[col].nunique()
        numeric_parts = [uni if val == 'binNA' else int(re.findall(r'\d+', val)[0]) for val in unique_bin_ranges]
        # numeric_parts = [int(re.findall(r'\d+', val)[0]) for val in unique_bin_ranges]
        unique_bin_ranges = unique_bin_ranges[np.argsort(numeric_parts)]
        df_new_cat = pd.DataFrame({"column_name": [col] * len(unique_bin_ranges), "bin ranges": unique_bin_ranges})
        df_new_cat = df_new_cat.merge(unique_values_in_bins.rename(columns={'bin': 'bin ranges', col: 'values in bin'}))
        # df_new_cat = df_new_cat.merge((binned_df_1[col].value_counts() / len(binned_df_1) * 100).reset_index().rename(columns={'index': 'bin ranges', col: 'count%'}).sort_values(by='bin ranges').reset_index(drop=True), on='bin ranges').round(2)
        df_new_cat = df_new_cat.merge((binned_df_1[col].value_counts() / len(binned_df_1) * 100).reset_index().rename(columns={col: 'bin ranges', 'count': 'count%'}).sort_values(by='bin ranges').reset_index(drop=True), on='bin ranges').round(2)
        if self.__dependant_target_variable is not None:
            df_new_cat = df_new_cat.merge(binned_df_1.groupby(col).target.sum(numeric_only=True).reset_index().rename(columns={col: 'bin ranges', 'target': 'Event'}), on='bin ranges')
            df_new_cat = df_new_cat.merge(binned_df_1.groupby(col).target.mean(numeric_only=True).reset_index().rename(columns={col: 'bin ranges', 'target': 'Mean_DV'}), on='bin ranges')
            df_new_cat['Index'] = (100 * df_new_cat['Mean_DV'] / binned_df_1.target.mean()).round()
        return df_new_cat

    def __write_summary_and_charts_to_excel(self, binned_df, sumar, binned_df_1, wb, ws, to_be_str, bb_no, plot_cell, cond_format, header_vales=None, for_groupby=None, dep_target_var=None):
        """
        Creates an Excel file with the generated summary of bins along with their charts

        :param binned_df: Dataframe containing the binned data
        :param sumar: Function to generate the summary of the bins
        :param binned_df_1: Dataframe containing the binned data with the target column
        :param wb: Work book
        :param ws: Work sheet
        :param to_be_str: The column which to be represented as string
        :param bb_no: Until which column the bottom border should apply
        :param plot_cell: Column from which the plotting of the charts starts
        :param header_vales: Values of the headers
        :param for_groupby: Dataframe which is used to find the values in a bin
        """
        row_num = 1

        headers = header_vales
        ws.append(headers)
        fill_color = PatternFill(start_color='89CFF0', end_color='89CFF0', fill_type='solid')
        if dep_target_var is not None:
            for i in range(1, 8):
                ws.cell(row=1, column=i).border = openpyxl.styles.borders.Border(bottom=openpyxl.styles.borders.Side(style='thick'))
                ws.cell(row=1, column=i).fill = fill_color
        else:
            for i in range(1, 5):
                ws.cell(row=1, column=i).border = openpyxl.styles.borders.Border(bottom=openpyxl.styles.borders.Side(style='thick'))
                ws.cell(row=1, column=i).fill = fill_color

        for col in binned_df.columns:
            # add the bin summary of a column to the Excel
            df_new = sumar(for_groupby, col, binned_df_1)
            df_new[to_be_str] = df_new[to_be_str].astype(str)
            start_row = row_num + 1
            for row in df_new.iterrows():
                row_num += 1
                for col_num, value in enumerate(row[1]):
                    ws.cell(row=row_num, column=col_num + 1, value=value)
                    ws.row_dimensions[row_num].height = 40
            end_row = row_num

            # border between different column variable
            for i in range(1, bb_no):
                ws.cell(row=row_num, column=i).border = openpyxl.styles.borders.Border(bottom=openpyxl.styles.borders.Side(style='thick'))

            # create charts for each column and set their metrics
            if dep_target_var is not None:
                combo_chart = BarChart()
                combo_chart.y_axis.majorGridlines = None
                line_chart = LineChart()
                line_chart.y_axis.majorGridlines = None
                combo_chart.y_axis.title = 'Count%'
                combo_chart.x_axis.title = 'Bin Ranges'
                line_chart.y_axis.title = 'Index'
                combo_chart.title = 'Distribution and Index of ' + col
                combo_chart.style = 10
                no_of_rows = end_row - start_row + 1
                combo_chart.height = min(1.5 * no_of_rows, 10)
                combo_chart.width = 10
                line_chart.y_axis.crosses = "max"
                line_chart.y_axis.axId = 200

                value_counts = Reference(ws, min_col=4, min_row=row_num - len(df_new), max_row=row_num)
                bin_ranges = Reference(ws, min_col=2, min_row=row_num - len(df_new) + 1, max_row=row_num)
                index = Reference(ws, min_col=7, min_row=row_num - len(df_new), max_row=row_num)
                combo_chart.add_data(value_counts, titles_from_data=True)
                combo_chart.set_categories(bin_ranges)
                line_chart.add_data(index, titles_from_data=True)
                line_chart.set_categories(bin_ranges)

                #     this code is used to make the maximum of the both y-axes same
                #     combo_chart.y_axis.scaling.min = 0
                #     line_chart.y_axis.scaling.min = 0
                #     max_value_counts = max(df_new['value_count'])
                #     max_index = max(df_new['Index'])
                #     max_value = max(max_value_counts, max_index)
                #     combo_chart.y_axis.scaling.max = max_value
                #     line_chart.y_axis.scaling.max = max_value

                combo_chart += line_chart
                combo_chart.legend = None
            else:
                combo_chart = BarChart()
                combo_chart.y_axis.majorGridlines = None
                combo_chart.y_axis.title = 'Count%'
                combo_chart.x_axis.title = 'Bin Ranges'
                combo_chart.title = 'Distribution and Index of ' + col
                combo_chart.style = 10
                no_of_rows = end_row - start_row + 1
                combo_chart.height = min(1.5 * no_of_rows, 10)
                combo_chart.width = 10

                value_counts = Reference(ws, min_col=4, min_row=row_num - len(df_new), max_row=row_num)
                bin_ranges = Reference(ws, min_col=2, min_row=row_num - len(df_new) + 1, max_row=row_num)
                combo_chart.add_data(value_counts, titles_from_data=True)
                combo_chart.set_categories(bin_ranges)

                combo_chart.legend = None

            # set the title size of the charts
            def set_chart_title_size(chart, size):
                paraprops = ParagraphProperties()
                paraprops.defRPr = CharacterProperties(sz=size)
                for para in chart.title.tx.rich.paragraphs:
                    para.pPr = paraprops
            set_chart_title_size(combo_chart, size=1000)

            ws.add_chart(combo_chart, plot_cell + str(start_row))

            #conditional formatting for the index column
            if dep_target_var is not None:
                x = ws.max_row
                rule = ColorScaleRule(start_type='min',
                                      start_color='00FF0000',
                                      mid_type='percentile',
                                      mid_value=50,
                                      mid_color='00FFFF00',  # Yellow
                                      end_type='max',
                                      end_color='0000FF00')
                colu = cond_format
                ws.conditional_formatting.add(f"{colu}2:{colu}{x + 1}", rule)
        if dep_target_var is not None:
            ws["P2"].value = "Index: Index tells you how a bin is performing compared to the entire dataset."
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 15
        wb.save("profiled_data.xlsx")

    def __ordinal_normal(self, df, column, ordinal_dict=None):
        """
        This function is used to replace the values of a dummy column to bin values (bin1, bin2...) based on the order given

        :param df: Dataframe
        :param column: Column Name
        :param ordinal_dict: Order of values in the column in the form of a dictionary
        :return: Binned column
        """
        if ordinal_dict is None:
            ordinal_dict = {val: i + 1 for i, val in enumerate(sorted(df[column].unique()))}
        else:
            ordinal_dict = {key: f'bin{val}' for key, val in ordinal_dict.items()}
        data = df.copy()
        data[column] = data[column].replace(ordinal_dict)
        return data

    def __apply_ordinal_normal(self, col, the_ordinal, df_ordinal):
        """
        this function is used to apply the ordinal_normal function

        :param col: column name
        :param the_ordinal: dictionary which contains the order of the values in the column
        :param df_ordinal: dataframe which contains only ordinal columns
        :return: ordinal_normal function
        """
        ordinal_dict = the_ordinal.loc[the_ordinal['variable'] == col, 'user action ordinal'].iloc[0]
        ordinal_dict = ast.literal_eval(ordinal_dict)
        return self.__ordinal_normal(df_ordinal, col, ordinal_dict)

    def __create_cat_binned_data(self, file_path=None, dependant_target_variable=None, no_of_bins=None, ordinal_binning='naive', max_thre=10, min_thre=5, tolerence=2, flag='ignore', min_cluster_size=0.05, max_clusters=10):
        """
        Calculated bins for each categorical column in a pandas Dataframe and their summary with plots

        :param file_path: string, path to file which contains either the column is nominal or ordinal
        :param dependant_target_variable: string, the name of the dependent variable column
        :param no_of_bins: int, the number of bins for grouping the data

        :return: an Excel file which contains the summary of each binned column and their respective charts
        """
        print('categorical binning start')
        df_org_cat = self.df.copy()
        if dependant_target_variable is not None:
            df_org_cat[dependant_target_variable] = df_org_cat[dependant_target_variable].astype('int64')
            df_org_cat = df_org_cat.rename(columns={dependant_target_variable: 'target'})
            dependant_target_variable = 'target'
        df_cat = df_org_cat.select_dtypes(include=['object'])
        # df_cat = pd.concat([df_cat, df_org_cat['target']], axis=1)

        # remove columns with only one unique values
        unique_counts = df_cat.nunique()
        unique_cols = unique_counts[unique_counts == 1].index.tolist()
        df_cat = df_cat.drop(unique_cols, axis=1)

        if file_path is None:
            # if file_path is not provided we will consider all categorical columns to be nominal
            df_nominal = df_cat.copy()

            if dependant_target_variable is not None:
                df_nominal = pd.concat([df_nominal, df_org_cat[dependant_target_variable]], axis=1)

                tqdm.pandas(dynamic_ncols=True, position=0)
                binned_df_nominal = df_nominal.progress_apply(lambda x: self.__woe_iv(df_nominal, x.name, dependant_target_variable, no_of_bins))
                binned_df_nominal.drop(dependant_target_variable, axis=1, inplace=True)
                binned_df_nominal = binned_df_nominal.applymap(lambda x: 'NA' if pd.isnull(x) else x)
                binned_df_nominal = binned_df_nominal.astype('category')

                cols_with_one_unique_bin = binned_df_nominal.columns[binned_df_nominal.nunique() == 1]
                binned_df_nominal.drop(cols_with_one_unique_bin, axis=1, inplace=True)

                binned_df_nominal_1 = pd.concat([binned_df_nominal, df_org_cat[dependant_target_variable]], axis=1)
            else:
                tqdm.pandas(dynamic_ncols=True, position=0)
                binned_df_nominal = df_nominal.progress_apply(lambda x: self.__naive_cat_bin(df_nominal, x.name, max_thre, min_thre, tolerence, flag))
                binned_df_nominal = binned_df_nominal.dropna(axis=1, how='all')
                binned_df_nominal = binned_df_nominal.astype('category')

                # for col in df_nominal.columns:
                #     self.naive_cat_bin(df_nominal, col, max_thre, min_thre, tolerence)
                # binned_df_nominal = df_nominal.copy()

                cols_with_one_unique_bin = binned_df_nominal.columns[binned_df_nominal.nunique() == 1]
                binned_df_nominal.drop(cols_with_one_unique_bin, axis=1, inplace=True)

                binned_df_nominal_1 = binned_df_nominal.copy()

            wb = openpyxl.load_workbook('profiled_data.xlsx')
            ws = wb.create_sheet()
            ws.title = 'Nominal Columns'

            if dependant_target_variable is not None:
                headers = ['column name', 'bin ranges', 'value in bin', 'count%', 'Event', 'Mean_DV', 'Index']
                self.__write_summary_and_charts_to_excel(binned_df_nominal, self.__col_bin_sumar_cat, binned_df_nominal_1, wb, ws, 'values in bin', 8, 'I', 'G', header_vales=headers, for_groupby=df_cat, dep_target_var=dependant_target_variable)
            else:
                headers = ['column name', 'bin ranges', 'value in bin', 'count%']
                self.__write_summary_and_charts_to_excel(binned_df_nominal, self.__col_bin_sumar_cat, binned_df_nominal_1, wb, ws, 'values in bin', 5, 'I', 'G', header_vales=headers, for_groupby=df_cat)

            print('categorical binning end')
            print('saving')

            wb.save('profiled_data.xlsx')

        else:
            # read the initial generated Excel file and separate ordinal columns from nominal columns
            df_excel = pd.read_excel(file_path)
            nominal_cols = df_excel[df_excel['user action type'] == 'nominal']['variable'].tolist()
            ordinal_cols = df_excel[df_excel['user action type'] == 'ordinal']['variable'].tolist()

            df_nominal = df_cat[df_cat.columns.intersection(nominal_cols)]
            df_ordinal = df_cat[df_cat.columns.intersection(ordinal_cols)]

            if dependant_target_variable is not None:
                df_nominal = pd.concat([df_nominal, df_org_cat[dependant_target_variable]], axis=1)

                tqdm.pandas(dynamic_ncols=True, position=0)
                binned_df_nominal = df_nominal.progress_apply(lambda x: self.__woe_iv(df_nominal, x.name, dependant_target_variable, no_of_bins))
                binned_df_nominal.drop(dependant_target_variable, axis=1, inplace=True)
                binned_df_nominal = binned_df_nominal.applymap(lambda x: 'NA' if pd.isnull(x) else x)
                binned_df_nominal = binned_df_nominal.astype('category')

                cols_with_one_unique_bin = binned_df_nominal.columns[binned_df_nominal.nunique() == 1]
                binned_df_nominal.drop(cols_with_one_unique_bin, axis=1, inplace=True)

                binned_df_nominal_1 = pd.concat([binned_df_nominal, df_org_cat[dependant_target_variable]], axis=1)
            else:
                tqdm.pandas(dynamic_ncols=True, position=0)
                binned_df_nominal = df_nominal.progress_apply(lambda x: self.__naive_cat_bin(df_nominal, x.name, max_thre, min_thre, tolerence, flag))
                binned_df_nominal = binned_df_nominal.dropna(axis=1, how='all')
                binned_df_nominal = binned_df_nominal.astype('category')

                # for col in df_nominal.columns:
                #     self.naive_cat_bin(df_nominal, col, max_thre, min_thre, tolerence)
                # binned_df_nominal = df_nominal.copy()

                cols_with_one_unique_bin = binned_df_nominal.columns[binned_df_nominal.nunique() == 1]
                binned_df_nominal.drop(cols_with_one_unique_bin, axis=1, inplace=True)

                binned_df_nominal_1 = binned_df_nominal.copy()

            wb = openpyxl.load_workbook('profiled_data.xlsx')
            ws = wb.create_sheet()
            ws.title = 'Nominal Columns'
            if dependant_target_variable is not None:
                headers = ['column name', 'bin ranges', 'value in bin', 'count%', 'Event', 'Mean_DV', 'Index']
                self.__write_summary_and_charts_to_excel(binned_df_nominal, self.__col_bin_sumar_cat, binned_df_nominal_1, wb, ws, 'values in bin', 8, 'I', 'G', header_vales=headers, for_groupby=df_cat, dep_target_var=dependant_target_variable)
            else:
                headers = ['column name', 'bin ranges', 'value in bin', 'count%']
                self.__write_summary_and_charts_to_excel(binned_df_nominal, self.__col_bin_sumar_cat, binned_df_nominal_1, wb, ws, 'values in bin', 5, 'I', 'G', header_vales=headers, for_groupby=df_cat)

            if not df_ordinal.empty:
                ordinal_columns = df_excel[df_excel.loc[:, 'user action type'] == 'ordinal']
                # the_ordinal contains column names and their corresponding dictionary orders
                the_ordinal = ordinal_columns.iloc[:, [0, 20]].reset_index(drop=True)
                binned_df_ordinal = df_ordinal.copy()

                if ordinal_binning == 'hierarchical':
                    binned_df_ordinal = df_ordinal.apply(lambda x: self.__hierarchical_clustering(df_ordinal, x.name, eval(the_ordinal.loc[the_ordinal.variable == x.name, 'user action ordinal'].values[0]), min_cluster_size=min_cluster_size, max_clusters=max_clusters))
                elif ordinal_binning == 'kprototype':
                    binned_df_ordinal = pd.concat([self.__kprototype(df_ordinal, col, eval(the_ordinal.loc[the_ordinal['variable'] == col, 'user action ordinal'].values[0]), max_clusters=max_clusters) for col in df_ordinal], axis=1).reset_index(drop=True)
                    # binned_df_ordinal = df_ordinal.apply(lambda x: self.__kprototype(df_ordinal, x.name, eval(the_ordinal.loc[the_ordinal.variable == x.name, 'user action ordinal'].values[0]), max_clusters=no_of_bins))
                elif ordinal_binning == 'no binning':
                    for col in binned_df_ordinal:
                        binned_df_ordinal = self.__apply_ordinal_normal(col, the_ordinal, binned_df_ordinal)
                else:
                    binned_df_ordinal = df_ordinal.apply(lambda x: self.__naive_ord_cat_bin(df_ordinal, x.name, eval(the_ordinal.loc[the_ordinal.variable == x.name, 'user action ordinal'].values[0]), max_thre, min_thre, tolerence))

                cols_with_one_unique_bin = binned_df_ordinal.columns[binned_df_ordinal.nunique() == 1]
                binned_df_ordinal.drop(cols_with_one_unique_bin, axis=1, inplace=True)

                if dependant_target_variable is not None:
                    binned_df_ordinal_1 = pd.concat([binned_df_ordinal, df_org_cat[dependant_target_variable]], axis=1)
                else:
                    binned_df_ordinal_1 = binned_df_ordinal.copy()

                ws1 = wb.create_sheet()
                ws1.title = 'Ordinal Columns'

                if dependant_target_variable is not None:
                    self.__write_summary_and_charts_to_excel(binned_df_ordinal, self.__col_bin_sumar_cat, binned_df_ordinal_1, wb, ws1, 'values in bin', 8, 'I', 'G', header_vales=headers, for_groupby=df_cat, dep_target_var=dependant_target_variable)
                else:
                    self.__write_summary_and_charts_to_excel(binned_df_ordinal, self.__col_bin_sumar_cat, binned_df_ordinal_1, wb, ws1, 'values in bin', 5, 'I', 'G', header_vales=headers, for_groupby=df_cat)

            print('categorical binning end')
            print('saving')

            wb1 = openpyxl.load_workbook(file_path)
            ws1 = wb1.worksheets[0]

            ws2 = wb.create_sheet()
            ws2.title = 'Summary'
            mr = ws1.max_row
            mc = ws1.max_column
            for i in range(1, mr + 1):
                for j in range(1, mc + 1):
                    c = ws1.cell(row=i, column=j)
                    ws2.cell(row=i, column=j).value = c.value

            red_fill = PatternFill(start_color='FFCCCB', end_color='FFCCCB', fill_type='solid')
            for row in ws2.iter_rows(min_row=2, min_col=2, max_col=2):
                for cell in row:
                    if cell.value is not None and cell.value > 80:
                        cell.fill = red_fill
                        ws2.cell(row=cell.row, column=1).fill = red_fill
            for row in ws2.iter_rows(min_row=2, min_col=4, max_col=4):
                for cell in row:
                    if cell.value is not None and cell.value > 10:
                        cell.fill = red_fill
                        ws2.cell(row=cell.row, column=1).fill = red_fill
            for row in ws2.iter_rows(min_row=2, min_col=9, max_col=9):
                for cell in row:
                    if cell.value is not None and (cell.value > 1 or cell.value < -1):
                        cell.fill = red_fill
                        ws2.cell(row=cell.row, column=1).fill = red_fill
            for col in ws2.columns:
                ws2.column_dimensions[col[0].column_letter].width = 20
            for row in ws2.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            wb.save('profiled_data.xlsx')

    def __vba_check(self, summary_path):
        """
        Creates a macro and assigns it to a button when clicked it ensures only the feature present in that sheets will be present in other sheets also

        :return:
        """
        if summary_path is not None:
            df_excel = pd.read_excel(summary_path)
            ordinal_cols = df_excel[df_excel['user action type'] == 'ordinal']['variable'].tolist()
            excel = win32.gencache.EnsureDispatch('Excel.Application')
            excel.Visible = True
            cwd = os.getcwd()
            cwd = cwd + '\profiled_data.xlsx'
            # wb = excel.Workbooks.Open(r'C:\Users\PrudhvitejaCherukuri\PycharmProjects\packagetest\profiled_data.xlsx')
            wb = excel.Workbooks.Open(r'' + cwd)
            sheet = wb.Worksheets('Summary')

            # Add a button to the worksheet using Shapes collection
            button = sheet.Buttons().Add(Left=3100, Top=0, Width=200, Height=30)
            button.Caption = 'Apply Filter'
            button.Name = 'RunMacro'
            button.OnAction = 'HideRows'

            # Define the macro
            if len(ordinal_cols) == 0:
                macro_name = \
                    '''
                    Sub HideRows()
                        Dim sheet1 As Worksheet
                        Dim sheet2 As Worksheet
                        Dim sheet3 As Worksheet
                        Dim lastRow1 As Long
                        Dim lastRow2 As Long
                        Dim lastRow3 As Long
                        Dim i As Long
                        Dim j As Long
                        Dim k AS Long
                        Dim foundMatch As Boolean

                        Set sheet1 = ThisWorkbook.Sheets("Numerical Columns")
                        Set sheet2 = ThisWorkbook.Sheets("Summary")
                        Set sheet3 = ThisWorkbook.Sheets("Nominal Columns")

                        sheet1.Rows.EntireRow.Hidden = False
                        sheet3.Rows.EntireRow.Hidden = False

                        lastRow1 = sheet1.Cells(Rows.Count, 1).End(xlUp).Row
                        lastRow2 = sheet2.Cells(Rows.Count, 1).End(xlUp).Row
                        lastRow3 = sheet3.Cells(Rows.Count, 1).End(XlUp).Row

                        For k = 2 To lastRow3
                            foundMatch2 = False
                            For Each cell In Sheet2.Range("A2:A" & lastRow2).SpecialCells(xlCellTypeVisible)
                                If sheet3.Cells(k, 1).Value = cell.Value Then
                                    foundMatch2 = True
                                    Exit for
                                End If
                            Next cell

                            If Not foundMatch2 Then
                                sheet3.Rows(k).Hidden = True
                            Else
                                sheet3.Rows(k).Hidden = False
                            End If
                        Next k

                        For i = 2 To lastRow1
                            foundMatch = False
                            For Each cell In sheet2.Range("A2:A" & lastRow2).SpecialCells(xlCellTypeVisible)
                                If sheet1.Cells(i, 1).Value = cell.Value Then
                                    foundMatch = True
                                    Exit For
                                End If
                            Next cell

                            If Not foundMatch Then
                                sheet1.Rows(i).Hidden = True
                            Else
                                sheet1.Rows(i).Hidden = False
                            End If
                        Next i

                    End Sub
                    '''
            else:
                macro_name = \
                    '''
                    Sub HideRows()
                        Dim sheet1 As Worksheet
                        Dim sheet2 As Worksheet
                        Dim sheet3 As Worksheet
                        Dim sheet4 As Worksheet
                        Dim lastRow1 As Long
                        Dim lastRow2 As Long
                        Dim lastRow3 As Long
                        Dim lastRow4 As Long
                        Dim i As Long
                        Dim j As Long
                        Dim k AS Long
                        Dim l As Long
                        Dim foundMatch As Boolean
        
                        Set sheet1 = ThisWorkbook.Sheets("Numerical Columns")
                        Set sheet2 = ThisWorkbook.Sheets("Summary")
                        Set sheet3 = ThisWorkbook.Sheets("Nominal Columns")
                        Set Sheet4 = ThisWorkbook.Sheets("Ordinal Columns")
        
                        sheet1.Rows.EntireRow.Hidden = False
                        sheet3.Rows.EntireRow.Hidden = False
                        sheet4.Rows.EntireRow.Hidden = False
        
                        lastRow1 = sheet1.Cells(Rows.Count, 1).End(xlUp).Row
                        lastRow2 = sheet2.Cells(Rows.Count, 1).End(xlUp).Row
                        lastRow3 = sheet3.Cells(Rows.Count, 1).End(XlUp).Row
                        lastRow4 = Sheet4.Cells(Rows.Count, 1).End(XlUp).Row
        
                        For k = 2 To lastRow3
                            foundMatch2 = False
                            For Each cell In Sheet2.Range("A2:A" & lastRow2).SpecialCells(xlCellTypeVisible)
                                If sheet3.Cells(k, 1).Value = cell.Value Then
                                    foundMatch2 = True
                                    Exit for
                                End If
                            Next cell
        
                            If Not foundMatch2 Then
                                sheet3.Rows(k).Hidden = True
                            Else
                                sheet3.Rows(k).Hidden = False
                            End If
                        Next k
        
                        For i = 2 To lastRow1
                            foundMatch = False
                            For Each cell In sheet2.Range("A2:A" & lastRow2).SpecialCells(xlCellTypeVisible)
                                If sheet1.Cells(i, 1).Value = cell.Value Then
                                    foundMatch = True
                                    Exit For
                                End If
                            Next cell
        
                            If Not foundMatch Then
                                sheet1.Rows(i).Hidden = True
                            Else
                                sheet1.Rows(i).Hidden = False
                            End If
                        Next i
                        
                        For i = 2 To lastRow4
                            foundMatch3 = False
                            For Each cell In sheet2.Range("A2:A" & lastRow2).SpecialCells(xlCellTypeVisible)
                                If sheet4.Cells(i, 1).Value = cell.Value Then
                                    foundMatch3 = True
                                    Exit For
                                End If
                            Next cell
        
                            If Not foundMatch Then
                                sheet4.Rows(i).Hidden = True
                            Else
                                sheet4.Rows(i).Hidden = False
                            End If
                        Next i
        
                    End Sub
                    '''

            # Add the macro to the workbook
            xlmodule = wb.VBProject.VBComponents.Add(1)
            xlmodule.CodeModule.AddFromString(macro_name.strip())

            wb.Save()
            # excel.Quit()

    def variable_profiling(self, summary_path=None, numerical_binning=None, dep_var=None, depth=None, no_of_bins=None, ordinal_binning_type='kprototype', max_threshold=10, min_threshold=5, tolerence=2, flag='ignore', min_cluster_size=0.05, max_clusters=10):
        """
        This function calls all the required function to perform variable profiling

        :param flag: Flag
        :param tolerence: Tolerance
        :param min_threshold: Minimum threshold
        :param max_threshold: Maximum threshold
        :param ordinal_binning_type: Type of binning function to perform on ordinal categorical column
        :param summary_path: Generated summary file
        :param numerical_binning: Type of binning to perform numerical columns
        :param dep_var: Dependant target variable
        :param depth: Depth of tree in case of tree based binning
        :param no_of_bins: Maximum  number of bins allowed
        """
        self.__transform(treatment_file=summary_path)
        self.__create_binned_data(func=numerical_binning, dep_var=dep_var, depth=depth, no_of_bins=no_of_bins)
        self.__create_cat_binned_data(file_path=summary_path, dependant_target_variable=dep_var, no_of_bins=no_of_bins, ordinal_binning=ordinal_binning_type, max_thre=max_threshold, min_thre=min_threshold, tolerence=tolerence, flag=flag, min_cluster_size=min_cluster_size, max_clusters=max_clusters)
        self.__vba_check(summary_path)
        print('Saved')

